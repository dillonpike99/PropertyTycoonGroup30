from openpyxl import load_workbook

class Parser:

    def __init__(self):
        boardFile = load_workbook("Data/PropertyTycoonBoardData.xlsx")
        self.boardSheet = boardFile["Sheet1"]
        cardsFile = load_workbook("Data/PropertyTycoonCardData.xlsx")
        self.cardsSheet = cardsFile["Sheet1"]

    def getTiles(self):
        columns = ["A", "B", "D", "F", "H", "I", "K", "L", "M", "N", "O"]
        tiles = []
        for i in range(5, 45):
            values = []
            for j in columns:
                values.append(self.boardSheet[f"{j}{i}"].value)
                if j == "F" and values[-1] == "No":
                    break
                elif j == "H":
                    if values[2] in ["Station", "Utilities"]:
                        break
            tiles.append(values)
        return tiles

    def getCards(self):
        potLuck = []
        opportunityKnocks = []
        for i in range(6, 23):
            potLuck.append(self.cardsSheet[f"A{i}"].value[1:-1])
        for i in range(27, 43):
            opportunityKnocks.append(self.cardsSheet[f"A{i}"].value[1:-1])
        return (potLuck, opportunityKnocks)